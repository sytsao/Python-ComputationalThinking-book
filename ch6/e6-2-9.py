#e6-2-9 使用  openpyxl  讀寫 excel 檔案
from openpyxl import load_workbook
wb = load_workbook('data7.xlsx') 
ws1 = wb.active 
rows=ws1.max_row
cols=ws1.max_column
print('資料總列數 =',rows)
print('資料總行數 =',cols)
ws2 = wb.create_sheet()
ws2.title = '成績計算'
ws2['A1'] ='學號'
ws2['B1'] ='國文'
ws2['C1'] ='英文'
ws2['D1'] ='數學'
ws2['E1'] ='總分'
for i in range(2,rows+2):
    sum=0
    for j in range(2, cols+1):
       sum+=ws1.cell(row=i-1, column=j).value
    ws2.cell(row=i,column=1).value = ws1.cell(row=i-1, column=1).value
    ws2.cell(row=i,column=2).value = ws1.cell(row=i-1, column=2).value
    ws2.cell(row=i,column=3).value = ws1.cell(row=i-1, column=3).value
    ws2.cell(row=i,column=4).value = ws1.cell(row=i-1, column=4).value
    ws2.cell(row=i,column=5).value = sum
wb.save('data7_out.xlsx')
